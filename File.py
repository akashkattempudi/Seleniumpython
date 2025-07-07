import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl
def product(path,colname,productname,newvalue):
    a = openpyxl.load_workbook(path)
    b = a.active
    d = {}
    for i in range(1, b.max_column + 1):
        if b.cell(row=1, column=i).value == colname:
            d["col"] = i
    for i in range(1, b.max_row + 1):
        for j in range(1, b.max_column + 1):
            if b.cell(row=i, column=j).value == productname:
                d["row"] = i
    b.cell(row=d["row"], column=d["col"]).value = newvalue
    a.save(path)
a = webdriver.Chrome()
a.implicitly_wait(5)
a.get("https://rahulshettyacademy.com/upload-download-test/index.html")
path = "C:\\Users\\akash\\Downloads\\download (1).xlsx"
# a.find_element(By.CSS_SELECTOR,"#downloadButton").click()
product(path,"price","Apple",243)
c = a.find_element(By.CSS_SELECTOR,"#fileinput")
c.send_keys(path)
z = (By.CSS_SELECTOR,"div[role='alert']")
b = WebDriverWait(a,5)
b.until(expected_conditions.visibility_of_element_located(z))
print(a.find_element(*z).text)
time.sleep(4)